import re
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import json
import openpyxl
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from ctypes import *


def read_user_info(filename: str):
    file = open(filename, "rb")
    file_json = json.load(file)
    username = file_json["username"]
    password = file_json["password"]
    file.close()
    return username, password


def change(chooses: str) -> str:
    dic = {'A': 1, 'B': 2, 'C': 3, 'D': 4, 'E': 5, 'F': 6}
    if chooses == "正确":
        return "1"
    if chooses == "错误":
        return "2"
    mat = re.match(r'[A-F]', chooses)
    if mat and len(chooses) == 1:
        return str(dic[chooses])
    mat = re.match(r'[A-F]{2,}', chooses)
    temp_string = ""
    if mat:
        for i in chooses:
            if not dic.get(i):
                return "12"
            temp_string = temp_string + str(dic[i])
        return temp_string


class DATA:
    def __init__(self):
        try:
            self.book = openpyxl.load_workbook("question.xlsx")
            self.sheet_single = self.book.worksheets[0]
            self.sheet_multiple = self.book.worksheets[1]
            self.sheet_judgment = self.book.worksheets[2]
        except Exception:
            print(Exception)
            exit(9)

    def __del__(self):
        self.book.close()


class OPERATOR:
    def __init__(self):
        self.current_traversal_row = None
        self.is_find = False
        self.question_type = None
        self.current = None  # 截取（1/10）字符串
        self.question = None
        self.all_answer_list = None
        self.answer = None
        self.answer_col = None
        self.current_sheet = None
        self.rows = None
        self.columns = None
        self.dll = cdll.LoadLibrary("./function.dll")
        self.change_func = self.dll.func
        self.change_func.argtype = c_char_p
        self.change_func.restype = c_char_p
        self.data = DATA()  # 题库
        # set user-agent
        self.operations = webdriver.ChromeOptions()
        self.operations.add_argument(
            'User-Agent: Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
            'Chrome/95.0.4638.54 Safari/537.36')
        # use operation
        self.driver = webdriver.Chrome(options=self.operations)
        # 填写webdriver的保存目录
        self.driver = webdriver.Chrome()
        self.wait = WebDriverWait(self.driver, 5)

    def __del__(self):
        del self.data

    def is_element_exist(self, by: str, value: str) -> bool:
        try:
            self.driver.find_element(by, value)
            return True
        except:
            return False

    def get_cookies(self, username: str, password: str):
        # 记得写完整的url 包括http和https
        self.driver.get('http://59.63.212.72:9117/html/login.html')
        self.driver.find_element(By.ID, 'username').send_keys(username)
        self.driver.find_element(By.ID, 'password').send_keys(password)
        # 程序打开网页后20秒内 “手动登陆账户”
        time.sleep(20)
        with open('cookies.txt', 'w') as file:
            # 将cookies保存为json格式
            file.write(json.dumps(self.driver.get_cookies()))
            file.close()
        self.driver.quit()  # 关闭当前窗口 driver.quite（） driver.closed() 不同

    def op(self):
        self.driver.get('http://59.63.212.72:9117/index.html?t=undefined&n=0')
        self.driver.delete_all_cookies()
        with open('cookies.txt', 'r') as f:
            # 使用json读取cookies 注意读取的是文件 所以用load而不是loads
            cookies_list = json.load(f)
            for cookie in cookies_list:
                self.driver.add_cookie(cookie)
        self.driver.refresh()  # 刷新
        time.sleep(3)

        if self.wait.until(EC.presence_of_element_located((By.LINK_TEXT, "闯关练兵"))):
            self.driver.find_element(By.LINK_TEXT, "闯关练兵").click()
        if self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".span1-6:nth-child(2) .item-go"))):
            self.driver.find_element(By.CSS_SELECTOR, ".span1-6:nth-child(2) .item-go").click()
        time.sleep(3)
        self.choose_level()  # 选关

    def do_work(self):
        self.click_again()
        self.get_question_data()
        time.sleep(0.2)
        self.before_search()
        self.search()
        self.click_answers()
        print("\n")
        if self.is_element_exist(By.CSS_SELECTOR, ".button > span"):
            self.driver.find_element(By.CSS_SELECTOR, ".button > span").click()
        time.sleep(0.2)
        self.click_again()

    def click_again(self):
        # 点击重新闯关
        if self.driver.find_element(By.XPATH, "//div[2]/div[3]/ul/li[2]").is_displayed():
            self.driver.find_element(By.XPATH, "//div[2]/div[3]/ul/li[2]").click()

    def click_answers(self):
        if self.is_find:  # 找到问题
            str_abc = self.current_sheet.cell(row=self.current_traversal_row, column=self.answer_col).value
            str_abc = str(str_abc).encode("utf-8")
            self.answer = self.change_func(str_abc).decode("utf-8")  # 调用C++写的func
            answer_list = list(self.answer)
            for i in range(0, len(answer_list)):
                if self.is_element_exist(By.CSS_SELECTOR, "li:nth-child(%d) > img" % int(answer_list[i])):
                    self.driver.find_element(By.CSS_SELECTOR, "li:nth-child(%d) > img" % int(answer_list[i])).click()
        else:
            print("没找到答案")
            if self.is_element_exist(By.CSS_SELECTOR, "li:nth-child(1) > img"):
                self.driver.find_element(By.CSS_SELECTOR, "li:nth-child(1) > img").click()

    def search(self):  # 对题库搜索
        for i in range(1, self.rows + 1):
            if self.current_sheet.cell(row=i, column=2).value == self.question:
                print(self.current_sheet.cell(row=i, column=self.answer_col).value)
                self.is_find = True
                self.current_traversal_row = i
                break

    def before_search(self):
        self.current_sheet = self.data.sheet_single  # 给个默认值 提高健壮性
        self.columns = self.current_sheet.max_column  # 给个默认值 提高健壮性

        if self.question_type == "单选题":
            self.current_sheet = self.data.sheet_single
            self.columns = self.current_sheet.max_column
            for i in range(1, self.columns + 1):
                if self.current_sheet.cell(row=1, column=i).value == "正确答案":
                    self.answer_col = i
                    break

        if self.question_type == "多选题":
            self.current_sheet = self.data.sheet_multiple
            self.columns = self.current_sheet.max_column
            for i in range(1, self.columns + 1):
                if self.current_sheet.cell(row=1, column=i).value == "正确答案":
                    self.answer_col = i
                    break

        if self.question_type == "判断题":
            self.current_sheet = self.data.sheet_judgment
            self.columns = self.current_sheet.max_column
            for i in range(1, self.columns + 1):
                if self.current_sheet.cell(row=1, column=i).value == "正确答案":
                    self.answer_col = i
                    break

        print(self.current + self.question_type)
        print(self.question)
        print(self.all_answer_list)
        self.rows = self.current_sheet.max_row
        if self.current_sheet is None:
            print("current_sheet is None")
            exit(10)

    def get_question_data(self):
        # one, two = ["aa", "bb"]  “(1/10)  单选题”
        if self.is_element_exist(By.CSS_SELECTOR, ".exam-right > .title"):
            self.current, self.question_type = \
                self.driver.find_element(By.CSS_SELECTOR, ".exam-right > .title").text.split("  ")
        if self.is_element_exist(By.CSS_SELECTOR, ".exam-right-subject>.title"):
            self.question = self.driver.find_element(By.CSS_SELECTOR, ".exam-right-subject>.title").text
        if self.is_element_exist(By.CSS_SELECTOR, ".exam-right-subject-list"):
            self.all_answer_list = \
                self.driver.find_element(By.CSS_SELECTOR, ".exam-right-subject-list").text.splitlines()

    def choose_level(self):
        if self.is_element_exist(By.CSS_SELECTOR, ".span1-6:nth-child(9) span"):
            self.driver.find_element(By.CSS_SELECTOR, ".span1-6:nth-child(9) span").click()  # 选关卡
        elif self.is_element_exist(By.CSS_SELECTOR, ".span1-6:nth-child(8) span"):
            self.driver.find_element(By.CSS_SELECTOR, ".span1-6:nth-child(8) span").click()  # 选关卡
        elif self.is_element_exist(By.CSS_SELECTOR, ".span1-6:nth-child(7) span"):
            self.driver.find_element(By.CSS_SELECTOR, ".span1-6:nth-child(7) span").click()  # 选关卡
        elif self.is_element_exist(By.CSS_SELECTOR, ".span1-6:nth-child(6) span"):
            self.driver.find_element(By.CSS_SELECTOR, ".span1-6:nth-child(6) span").click()  # 选关卡
        elif self.is_element_exist(By.CSS_SELECTOR, ".span1-6:nth-child(5) span"):
            self.driver.find_element(By.CSS_SELECTOR, ".span1-6:nth-child(5) span").click()  # 选关卡
        elif self.is_element_exist(By.CSS_SELECTOR, ".span1-6:nth-child(4) span"):
            self.driver.find_element(By.CSS_SELECTOR, ".span1-6:nth-child(4) span").click()  # 选关卡
        elif self.is_element_exist(By.CSS_SELECTOR, ".span1-6:nth-child(3) span"):
            self.driver.find_element(By.CSS_SELECTOR, ".span1-6:nth-child(3) span").click()  # 选关卡
        elif self.is_element_exist(By.CSS_SELECTOR, ".span1-6:nth-child(2) span"):
            self.driver.find_element(By.CSS_SELECTOR, ".span1-6:nth-child(2) span").click()  # 选关卡
        elif self.is_element_exist(By.CSS_SELECTOR, ".span1-6:nth-child(1) span"):
            self.driver.find_element(By.CSS_SELECTOR, ".span1-6:nth-child(1) span").click()  # 选关卡

    def find_right_answers(self, _question_type: str, _question: str, _all_snswer_list: list):
        if _question_type == "单选题":
            self.data.sheet_single.find(_question)

    def do_single_question(self):
        pass

    def do_multiple_question(self):
        pass

    def do_judgment_question(self):
        pass
